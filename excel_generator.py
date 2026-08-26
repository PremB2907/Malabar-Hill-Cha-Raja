import openpyxl
import re
import json
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font

def normalize_header(value, index):
    text = str(value or '').strip()
    return text if text else f"Column {index + 1}"

def normalize_key(value):
    return re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower()).strip('_')

def number_from_value(value):
    if isinstance(value, (int, float)):
        return value
    cleaned = str(value or '').replace('₹', '').replace(',', '').strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0

def get_derived_amount(data):
    keys = list(data.keys())
    preferred = [
        'amount', 'donation_amount', 'total_amount', 'total',
        'paid_amount', 'paid', 'price', 'value', 'donation'
    ]
    for wanted in preferred:
        matched_key = next((k for k in keys if normalize_key(k) == wanted), None)
        if matched_key:
            val = number_from_value(data[matched_key])
            if val != 0:
                return val
                
    for key in keys:
        normalized = normalize_key(key)
        if any(term in normalized for term in ['amount', 'donation', 'total', 'paid', 'price', 'value']):
            val = number_from_value(data[key])
            if val != 0:
                return val
    return 0.0

def get_derived_quantity(data):
    keys = list(data.keys())
    preferred = ['quantity', 'qty', 'count', 'number_of_tshirts', 'tshirt_quantity']
    for wanted in preferred:
        matched_key = next((k for k in keys if normalize_key(k) == wanted), None)
        if matched_key:
            val = int(number_from_value(data[matched_key]))
            return val if val > 0 else 1
    return 1

def parse_workbook(file_stream):
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    if not wb.worksheets:
        raise ValueError("The Excel file does not contain a worksheet.")
    
    ws = wb.worksheets[0]
    
    # Read headers
    columns = []
    seen = set()
    
    # Iterate over first row
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    if not first_row:
        raise ValueError("The first row must contain column names.")
        
    for idx, cell_val in enumerate(first_row):
        header = normalize_header(cell_val, idx)
        original = header
        suffix = 2
        while header.lower() in seen:
            header = f"{original} {suffix}"
            suffix += 1
        seen.add(header.lower())
        columns.append(header)
        
    rows = []
    for row_num in range(2, ws.max_row + 1):
        # Retrieve row cells
        row_cells = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        if not row_cells:
            continue
            
        has_value = False
        data = {}
        for col_idx, col_name in enumerate(columns):
            val = row_cells[col_idx] if col_idx < len(row_cells) else None
            # Formatting Date
            if hasattr(val, 'isoformat'):
                val = val.isoformat()
            
            if val is not None and str(val).strip() != '':
                has_value = True
                
            data[col_name] = val if val is not None else ''
            
        if has_value:
            rows.append({
                "row_number": row_num,
                "data": data,
                "amount": get_derived_amount(data),
                "quantity": get_derived_quantity(data)
            })
            
    return {"worksheet_name": ws.title, "columns": columns, "rows": rows}

def export_offline_sheets(sheets, get_rows_func):
    """
    Export offline spreadsheets metadata and rows.
    sheets: list of offline sheet metadata records
    get_rows_func: function to get rows by sheet id
    """
    wb = openpyxl.Workbook()
    
    if not sheets:
        ws = wb.active
        ws.title = "No Records"
        ws.append(["No offline records available."])
    else:
        # Remove default sheet
        wb.remove(wb.active)
        
        for idx, sheet in enumerate(sheets):
            # Sheet titles in Excel have 31 char limit, no special characters
            safe_name = re.sub(r'[\\/*?:\[\]]', ' ', str(sheet.get("sheet_name", "Records")))
            safe_name = safe_name.strip()[:31] if safe_name.strip() else "Records"
            
            name = safe_name
            suffix = 2
            while name in wb.sheetnames:
                # Truncate to leave space for suffix
                name = f"{safe_name[:27]} {suffix}"
                suffix += 1
                
            ws = wb.create_sheet(title=name)
            
            record_type_str = "Offline Donation" if sheet.get("record_type") == "donation" else "Offline T-Shirt"
            ws.append(["Record Type", record_type_str])
            ws.append(["Original File", sheet.get("original_filename", "")])
            ws.append([]) # empty separator row
            
            columns = sheet.get("columns", [])
            ws.append(columns)
            
            rows = get_rows_func(sheet["id"])
            for row in rows:
                row_data = row.get("data", {})
                ws.append([row_data.get(c, '') for c in columns])
                
            # Formatting
            ws.row_dimensions[4].font = Font(bold=True)
            # Freeze panes below headers
            ws.freeze_panes = "A5"
            
            # Simple column width adjustment
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def export_dbt_receipts(receipts):
    """Export DBT receipts worksheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DBT Receipts"
    
    headers = [
        'Reference ID', 'Donor Name', 'Phone', 'Email', 'Amount', 
        'Transaction Reference', 'Uploaded Receipt', 'Status', 'Submitted At'
    ]
    ws.append(headers)
    
    for r in receipts:
        ws.append([
            r.get('reference_id', ''),
            r.get('donor_name', ''),
            r.get('phone', ''),
            r.get('email', ''),
            float(r.get('amount', 0.0)),
            r.get('transaction_ref', ''),
            r.get('original_filename', ''),
            r.get('status', ''),
            r.get('created_at', '')
        ])
        
    # Formatting
    ws.row_dimensions[1].font = Font(bold=True)
    ws.freeze_panes = "A2"
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)
        
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def export_combined_records(donations, tshirts, offline_donations, offline_tshirts):
    """Export combined online & offline donations and T-shirt bookings."""
    wb = openpyxl.Workbook()
    
    # Sheet 1: All Donations
    ws_donations = wb.active
    ws_donations.title = "All Donations"
    ws_donations.append(['Source', 'Receipt / Row', 'Name', 'Phone', 'Amount', 'Date'])
    
    for d in donations:
        ws_donations.append([
            'Online', 
            d.get('receipt_no', ''), 
            d.get('donor_name', ''), 
            d.get('phone', ''), 
            float(d.get('amount', 0.0)), 
            d.get('created_at', '')
        ])
        
    for r in offline_donations:
        data = r.get('data', {})
        keys = list(data.keys())
        
        name_key = next((k for k in keys if re.match(r'^(name|donor_name|donor|full_name)$', str(k).strip(), re.IGNORECASE)), None)
        if not name_key and keys:
            name_key = keys[0]
            
        phone_key = next((k for k in keys if re.search(r'phone|mobile|contact', str(k), re.IGNORECASE)), None)
        
        name_val = data.get(name_key, '') if name_key else ''
        phone_val = data.get(phone_key, '') if phone_key else ''
        
        ws_donations.append([
            'Offline',
            f"{r.get('sheet_name', '')} / Row {r.get('row_number', '')}",
            name_val,
            phone_val,
            float(r.get('amount', 0.0)),
            ''
        ])

    # Sheet 2: All T-Shirt Bookings
    ws_tshirts = wb.create_sheet(title="All T-Shirt Bookings")
    ws_tshirts.append(['Source', 'Receipt / Row', 'Name', 'Phone', 'Quantity', 'Amount', 'Date'])
    
    for o in tshirts:
        ws_tshirts.append([
            'Online',
            o.get('receipt_no', ''),
            o.get('buyer_name', ''),
            o.get('phone', ''),
            int(o.get('quantity', 1)),
            float(o.get('total_amount', 0.0)),
            o.get('created_at', '')
        ])
        
    for r in offline_tshirts:
        data = r.get('data', {})
        keys = list(data.keys())
        
        name_key = next((k for k in keys if re.match(r'^(name|buyer_name|buyer|full_name)$', str(k).strip(), re.IGNORECASE)), None)
        if not name_key and keys:
            name_key = keys[0]
            
        phone_key = next((k for k in keys if re.search(r'phone|mobile|contact', str(k), re.IGNORECASE)), None)
        
        name_val = data.get(name_key, '') if name_key else ''
        phone_val = data.get(phone_key, '') if phone_key else ''
        
        ws_tshirts.append([
            'Offline',
            f"{r.get('sheet_name', '')} / Row {r.get('row_number', '')}",
            name_val,
            phone_val,
            int(r.get('quantity', 1)),
            float(r.get('amount', 0.0)),
            ''
        ])

    # Format both sheets
    for ws in [ws_donations, ws_tshirts]:
        ws.row_dimensions[1].font = Font(bold=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
